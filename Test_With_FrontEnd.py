import tempfile
import streamlit as st
import io
import pandas as pd
import openpyxl
from datetime import datetime
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import re
import glob
import os
from openpyxl import load_workbook


def extract_far_dates(ws):
    """
    Extract year-end date from cell A2 and period-end date from cell C1.
    """
    try:
        year_end_cell = ws['A2'].value
        period_cell = ws['C1'].value

        year_end_date = pd.to_datetime(str(year_end_cell).split("Year End -")[-1].strip(), dayfirst=True)
        period_end_date = None
        if period_cell:
            match = re.search(r'(January|February|March|April|May|June|July|August|September|October|November|December)[\' ](\d{2,4})', str(period_cell))
            if match:
                period_str = match.group(0).replace("'", " 20")
                period_end_date = pd.to_datetime("01 " + period_str, dayfirst=True) + pd.offsets.MonthEnd(0)
        return year_end_date, period_end_date
    except Exception as e:
        raise ValueError(f"Failed to extract FAR dates: {e}")

# Utility Functions
# Utility Functions
def safe_float(value, context="unknown"):
    """
    Safely convert a value to float, handling strings and None values.
    Used to prevent string subtraction errors in Excel cell operations.
    """
    if value is None:
        return 0.0
    
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0

def find_account_row(wsTrans, accountName):
    """
    Find the row index of an account in Account Transactions sheet.
    Returns None if account not found.
    """
    for rowIndex in range(1, wsTrans.max_row + 1):
        if wsTrans.cell(row=rowIndex, column=1).value == accountName:
            return rowIndex
    return None

def create_fresh_stream(original_content):
    """Create a fresh BytesIO stream from original content"""
    return BytesIO(original_content)

def extract_month_key(transaction_date, format_type="MMMM YYYY"):
    """
    Extract standardized month key from transaction date
    
    Args:
        transaction_date: Date value from Excel cell
        format_type: Format type - "MMMM YYYY" (e.g., "January 2024") or "MMM YYYY" (e.g., "Jan 2024")
    
    Returns:
        str: Formatted month key or empty string if invalid
    """
    if isinstance(transaction_date, datetime):
        if format_type == "MMM YYYY":
            return transaction_date.strftime("%b %Y")
        else:
            return transaction_date.strftime("%B %Y")
    elif transaction_date:
        try:
            dt = pd.to_datetime(transaction_date, errors='coerce')
            if pd.notna(dt):
                if format_type == "MMM YYYY":
                    return dt.strftime("%b %Y")
                else:
                    return dt.strftime("%B %Y")
            else:
                return str(transaction_date)[:7]
        except Exception:
            return str(transaction_date)[:7]
    else:
        return ""

def apply_tax_component_formatting(ws):
    start_row, end_row = 15, 26  # Reduced from 27 to 26 since we removed one row
    start_col, end_col = 9, 11  # I=9, J=10, K=11 (shifted from G-I to I-K)
    bold_font = Font(name='Gill Sans MT', size=12, bold=True)
    regular_font = Font(name='Gill Sans MT', size=12, bold=False)
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=start_row, column=col)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for row in range(start_row + 1, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = regular_font
            cell.alignment = right_align if col > start_col else left_align
            cell.border = thin_border
            cell.fill = PatternFill(fill_type=None)
    # Apply percentage formatting to the Corporation Tax rate row
    for c in [10, 11]:  # Columns J and K
        cell = ws.cell(row=22, column=c)
        cell.number_format = '0.00%'
        cell.font = bold_font
        cell.alignment = right_align
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 14

def format_summary_table(ws, start_row=15):
    start_col = 1
    max_row = ws.max_row
    max_col = ws.max_column
    last_row = start_row
    last_col = start_col
    for r in range(start_row, max_row+1):
        row_has_data = False
        for c in range(start_col, max_col+1):
            val = ws.cell(row=r, column=c).value
            if val is not None and str(val).strip() != "":
                row_has_data = True
                if c > last_col:
                    last_col = c
        if row_has_data:
            last_row = r
    if last_row < start_row or last_col < start_col:
        return
    font = Font(name="Gill Sans MT", size=12)
    bold_font = Font(name="Gill Sans MT", size=12, bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for r in range(start_row, last_row+1):
        for c in range(start_col, last_col+1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            if r == start_row:
                cell.font = bold_font
                cell.alignment = Alignment(horizontal="center", vertical="bottom")
            else:
                cell.font = font
                if isinstance(val, (int, float)):
                    cell.alignment = Alignment(horizontal="right", vertical="bottom")
                    if val is not None:
                        if val < 0:
                            cell.number_format = "#,##0;(#,##0)"
                        elif val == 0:
                            cell.number_format = "#,##0;-#,##0;-"
                        else:
                            cell.number_format = "#,##0"
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="bottom")
            cell.border = border
    for c in range(start_col, last_col+1):
        maxlen = 0
        for r in range(start_row, last_row+1):
            v = ws.cell(row=r, column=c).value
            vstr = str(v) if v is not None else ""
            if len(vstr) > maxlen:
                maxlen = len(vstr)
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = max(10, min(maxlen+2, 40))
    if hasattr(ws, 'sheet_view'):
        ws.sheet_view.showGridLines = False

def _month_sort_key(x):
    dt = pd.to_datetime(x, errors='coerce')
    if pd.notna(dt):
        return (0, dt)
    return (1, str(x))

class FormatProcessorBase:
    """Base class for all format processors with common functionality"""
    
    def __init__(self, target_sheet, transactions_sheet):
        self.ws = target_sheet
        self.wsTrans = transactions_sheet
        self.accountName = target_sheet.cell(row=4, column=1).value
        
    def find_account_row(self):
        """Find account row in transactions sheet"""
        return find_account_row(self.wsTrans, self.accountName)
    
    def get_opening_balance(self, accountRow, format_type="I_minus_H"):
        """Get opening balance with different calculation methods"""
        accountRow = accountRow + 1  # Move to next row for balance
        val_h_raw = self.wsTrans.cell(row=accountRow, column=8).value
        val_i_raw = self.wsTrans.cell(row=accountRow, column=9).value
        
        if format_type == "I_minus_H":
            return safe_float(val_i_raw, f"openingBalance col I row {accountRow}") - safe_float(val_h_raw, f"openingBalance col H row {accountRow}")
        elif format_type == "H_minus_I":
            return safe_float(val_h_raw, f"openingBalance col H row {accountRow}") - safe_float(val_i_raw, f"openingBalance col I row {accountRow}")
        else:
            return safe_float(val_i_raw, f"openingBalance col I row {accountRow}") - safe_float(val_h_raw, f"openingBalance col H row {accountRow}")
    
    def setup_summary_headers(self, headers, start_row=15):
        """Setup standard summary table headers"""
        for idx, header in enumerate(headers):
            cell = self.ws.cell(row=start_row, column=idx + 1, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
        return start_row + 1

# Main Processing Function (full business logic from your Test.py)
def process_far_file(file_content):
    """
    Main processing function for FAR files.
    Handles Excel file processing with multiple format types.
    """
    # Store original content for multiple uses - avoid stream closure issues
    original_content = file_content
    
    # Load Excel data with fresh streams and data_only=True to avoid image issues
    df = pd.read_excel(create_fresh_stream(original_content), sheet_name='FAR', engine='openpyxl')
    wb = openpyxl.load_workbook(create_fresh_stream(original_content), data_only=True, read_only=False)


    # Extract year_end_date and period_end_date from FAR sheet
    df_far_head = pd.read_excel(create_fresh_stream(original_content), sheet_name='FAR', header=None, nrows=5, engine='openpyxl')
    year_end_date, period_end_date = None, None
    for i in range(5):
        row_vals = df_far_head.iloc[i].astype(str).tolist()
        for val in row_vals:
            if 'year end' in val.lower():
                match = re.search(r'(\d{1,2} [A-Za-z]+ \d{4})', val)
                if match:
                    try:
                        year_end_date = datetime.strptime(match.group(1), '%d %B %Y')
                    except Exception:
                        pass
            if 'management accounts' in val.lower():
                match = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', val)
                if match:
                    try:
                        period_end_date = datetime.strptime(match.group(1), '%d/%m/%Y')
                    except Exception:
                        pass
                match2 = re.search(r'QE?\s*([A-Za-z]+)[\'’]?(\d{2})', val)
                if match2:
                    try:
                        month = match2.group(1)
                        year = int('20' + match2.group(2))
                        # Create date and move to end of month
                        temp_date = datetime.strptime(f'{month} {year}', '%b %Y')
                        period_end_date = temp_date.replace(day=1) + pd.offsets.MonthEnd(0)
                        period_end_date = period_end_date.to_pydatetime()
                    except Exception:
                        pass
                match3 = re.search(r'([A-Za-z]+)\s+(\d{4})', val)
                if match3:
                    try:
                        # Create date and move to end of month
                        temp_date = datetime.strptime(f'{match3.group(1)} {match3.group(2)}', '%B %Y')
                        period_end_date = temp_date.replace(day=1) + pd.offsets.MonthEnd(0)
                        period_end_date = period_end_date.to_pydatetime()
                    except Exception:
                        pass

    if not year_end_date:
        raise Exception("❌ Could not extract year-end date from FAR sheet.")
    if not period_end_date:
        raise Exception("❌ Could not extract period end date (management account date) from FAR sheet.")

    # 🗓️ Compute year boundaries
    fy_end = pd.to_datetime(year_end_date)
    fy_start = fy_end - pd.DateOffset(years=1) + pd.DateOffset(days=1)
    mgmt_acct_month = pd.to_datetime(period_end_date)

    # 📂 Setup file label and temp output folder
    output_base_name = "uploaded_file"
    output_folder = tempfile.mkdtemp()

    # 📘 Workbook is already loaded above - no need to reload


    # --- MODULE 1: Split Account Transactions into separate sheets ---
    if "Account Transactions" in wb.sheetnames:
        wsSource = wb["Account Transactions"]
        clientName = wsSource["A2"].value
        excludeAccounts = [
            "Freehold Property", "Leasehold Property", "Leasehold Property Depreciation", "Plant & Machinery", "Plant & Machinery Depreciation", "Bar & Kitchen Equipment", "Bar & Kitchen Equipment Depreciation", "Furniture & Fixtures", "Furniture & Fixtures Depreciation", "Motor Vehicles", "Motor Vehicles Depreciation", "Property Improvements", "Property Improvements Depreciation", "Refurbishment", "Refurbishment Depreciation", "Goodwill", "Goodwill Amortisation", "Historical Adjustment"
        ]
        lastRow = wsSource.max_row
        headerRow = None
        for row in range(1, lastRow+1):
            if str(wsSource.cell(row=row, column=1).value).strip().upper() == "DATE":
                headerRow = row
                break
        transactionStart = None
        accountName = None
        for row in range(1, lastRow+1):
            val = wsSource.cell(row=row, column=1).value
            if val is None or str(val).strip() == "":
                if row+1 <= lastRow and wsSource.cell(row=row+1, column=1).value not in [None, ""]:
                    accountName = wsSource.cell(row=row+1, column=1).value
                    if accountName in excludeAccounts:
                        continue
                    def get_valid_sheet_name(name):
                        for ch in "/\\:?*[]":
                            name = name.replace(ch, "_")
                        return name[:31]
                    newSheetName = get_valid_sheet_name(accountName)
                    if newSheetName in wb.sheetnames:
                        wsNew = wb[newSheetName]
                    else:
                        wsNew = wb.create_sheet(title=newSheetName)
                    
                    # Remove gridlines from the new sheet
                    wsNew.sheet_view.showGridLines = False
                    
                    # PART 1: ROWS 1-4
                    wsNew["A1"] = clientName
                    wsNew["A2"] = f"Year End - {year_end_date.strftime('%d %B %Y')}"
                    wsNew["A3"] = f"Management Accounts : {period_end_date.strftime('%b\'%y').capitalize()}"
                    wsNew["A4"] = accountName
                    for r in range(1, 5):
                        cell = wsNew.cell(row=r, column=1)
                        cell.font = openpyxl.styles.Font(name="Gill Sans MT", size=12, bold=True)
                        cell.alignment = openpyxl.styles.Alignment(horizontal="left", vertical="bottom")
                    # PART 2: SUMMARY TABLE
                    wsNew["A6"] = "Date"
                    wsNew["B6"] = "Details"
                    wsNew["C6"] = "Amount £"
                    for c in range(1, 4):
                        cell = wsNew.cell(row=6, column=c)
                        cell.font = openpyxl.styles.Font(name="Gill Sans MT", size=12, bold=True)
                    wsNew["A7"].value = None
                    wsNew["B7"].value = None
                    wsNew["C7"].value = None
                    wsNew["A8"] = period_end_date.strftime('%d-%m-%Y')
                    wsNew["A8"].number_format = "dd-mm-yyyy"
                    wsNew["B8"] = accountName
                    wsNew["C8"] = None
                    # Apply font formatting to row 8
                    for c in range(1, 4):
                        cell = wsNew.cell(row=8, column=c)
                        cell.font = openpyxl.styles.Font(name="Gill Sans MT", size=12)
                    wsNew["A9"].value = None
                    wsNew["B9"].value = None
                    wsNew["C9"].value = None
                    # Add Total row at row 10
                    wsNew["A10"] = "Total"
                    wsNew["B10"].value = None
                    wsNew["C10"] = "=C8"  # Reference C8 to show same value as summary
                    border = openpyxl.styles.Border(
                        top=openpyxl.styles.Side(style='thin'),
                        bottom=openpyxl.styles.Side(style='thin')
                    )
                    for c in range(1, 4):
                        wsNew.cell(row=6, column=c).border = border
                        wsNew.cell(row=10, column=c).border = border
                        # Make Total row bold
                        wsNew.cell(row=10, column=c).font = openpyxl.styles.Font(name="Gill Sans MT", size=12, bold=True)
                    # Apply number formatting to C10 to match C8
                    wsNew["C10"].number_format = "#,##0.00;(#,##0.00)"

    # --- MODULE 2: Process Mappings sheet and apply FormatN logic ---
    if 'Mappings' in wb.sheetnames:
        ws_mappings = wb['Mappings']
        mappings = []
        for row in ws_mappings.iter_rows(min_row=2, max_row=ws_mappings.max_row, min_col=1, max_col=2, values_only=True):
            if row[0] and row[1]:
                mappings.append((str(row[0]), str(row[1])))
        for itemName, functionName in mappings:
            target_sheet = None
            for sheet_name in wb.sheetnames:
                if sheet_name not in ['Mappings', 'Account Transactions', 'P&L']:
                    ws_account = wb[sheet_name]
                    accountName = ws_account['A4'].value
                if accountName == itemName:
                    target_sheet = ws_account
                    break
            if target_sheet is not None:
                fmt = functionName.lower()
                # --- Format1 ---
                if fmt == 'format1':
                    processor = FormatProcessorBase(target_sheet, wb["Account Transactions"])
                    
                    # Find account row
                    accountRow = processor.find_account_row() 
                    if accountRow is None:
                        print(f"Account not found for Format1: {processor.accountName}")
                        continue
                    
                    # Calculate opening balance (I - H)
                    openingBalance = processor.get_opening_balance(accountRow, "I_minus_H")
                    
                    # Initialize data structures
                    accountDict = {}
                    uniqueMonths = set()
                    
                    # Process transactions starting from accountRow + 2
                    for rowIndex in range(accountRow + 2, processor.wsTrans.max_row + 1):
                        val = processor.wsTrans.cell(row=rowIndex, column=1).value
                        if val in ("Total", "Closing Balance"):
                            break
                        
                        # Extract account name from Column R (18), starting at 8th character
                        acc_name_val = processor.wsTrans.cell(row=rowIndex, column=18).value
                        if acc_name_val:
                            currentAccountName = str(acc_name_val)[7:] if len(str(acc_name_val)) >= 8 else str(acc_name_val)
                        else:
                            currentAccountName = ""
                        
                        # Extract transaction date and get month key
                        transactionDate = processor.wsTrans.cell(row=rowIndex, column=1).value
                        monthKey = extract_month_key(transactionDate, format_type="MMM YYYY")
                        if not monthKey:
                            continue
                        
                        # Add to unique months set
                        uniqueMonths.add(monthKey)
                        
                        # Initialize account entry if not exists
                        if currentAccountName not in accountDict:
                            accountDict[currentAccountName] = {}
                        if monthKey not in accountDict[currentAccountName]:
                            accountDict[currentAccountName][monthKey] = 0
                        
                        # Calculate transaction value (I - H)
                        val_i_raw = processor.wsTrans.cell(row=rowIndex, column=9).value
                        val_h_raw = processor.wsTrans.cell(row=rowIndex, column=8).value
                        val_i = safe_float(val_i_raw, f"format1 val_i row {rowIndex}")
                        val_h = safe_float(val_h_raw, f"format1 val_h row {rowIndex}")
                        accountDict[currentAccountName][monthKey] += val_i - val_h
                    
                    # Generate the summary table
                    summaryStartRow = 15
                    processor.ws.cell(row=summaryStartRow, column=1).value = "Account Name"
                    processor.ws.cell(row=summaryStartRow, column=2).value = "Opening Balance"
                    
                    # Sort months chronologically and add headers
                    sortedMonths = sorted(uniqueMonths, key=_month_sort_key)
                    for idx, month in enumerate(sortedMonths):
                        processor.ws.cell(row=summaryStartRow, column=3 + idx).value = month
                    processor.ws.cell(row=summaryStartRow, column=3 + len(sortedMonths)).value = "Closing Balance"
                    
                    # Fill in the summary data for each account
                    summaryRow = summaryStartRow + 1
                    for currentAccountName in accountDict.keys():
                        processor.ws.cell(row=summaryRow, column=1).value = currentAccountName
                        processor.ws.cell(row=summaryRow, column=2).value = None  # Opening balance per account not tracked
                        for idx, month in enumerate(sortedMonths):
                            processor.ws.cell(row=summaryRow, column=3 + idx).value = accountDict[currentAccountName].get(month, 0)
                        summaryRow += 1
                    
                    # Add the total row
                    processor.ws.cell(row=summaryRow, column=1).value = "Total"
                    processor.ws.cell(row=summaryRow, column=2).value = openingBalance
                    
                    # Calculate monthly totals and final closing balance
                    totalClosingBalance = openingBalance
                    for idx, month in enumerate(sortedMonths):
                        totalSum = sum(accountDict[acc].get(month, 0) for acc in accountDict)
                        processor.ws.cell(row=summaryRow, column=3 + idx).value = totalSum
                        totalClosingBalance += totalSum
                    
                    # Round the final closing balance to avoid floating-point precision issues
                    totalClosingBalance = round(totalClosingBalance, 2)
                    
                    # Set final closing balance
                    processor.ws.cell(row=summaryRow, column=3 + len(sortedMonths)).value = totalClosingBalance
                    
                    # Apply bold formatting to headers and total row
                    for col in range(1, 4 + len(sortedMonths)):
                        processor.ws.cell(row=summaryStartRow, column=col).font = openpyxl.styles.Font(bold=True)
                        processor.ws.cell(row=summaryRow, column=col).font = openpyxl.styles.Font(bold=True)
                    
                    # Apply summary table formatting
                    format_summary_table(processor.ws, start_row=15)
                # --- Format2 ---
                elif fmt == 'format2':
                    processor = FormatProcessorBase(target_sheet, wb["Account Transactions"])
                    
                    # Setup headers from row 5 of Account Transactions (preserving original headers)
                    summaryStartRow = 15
                    processor.ws.cell(row=summaryStartRow, column=1).value = processor.wsTrans.cell(row=5, column=1).value
                    processor.ws.cell(row=summaryStartRow, column=2).value = processor.wsTrans.cell(row=5, column=2).value
                    processor.ws.cell(row=summaryStartRow, column=3).value = processor.wsTrans.cell(row=5, column=5).value
                    processor.ws.cell(row=summaryStartRow, column=4).value = processor.wsTrans.cell(row=5, column=7).value
                    processor.ws.cell(row=summaryStartRow, column=5).value = processor.wsTrans.cell(row=5, column=8).value
                    processor.ws.cell(row=summaryStartRow, column=6).value = processor.wsTrans.cell(row=5, column=9).value
                    
                    # Make headers bold
                    for col in range(1, 7):
                        processor.ws.cell(row=summaryStartRow, column=col).font = openpyxl.styles.Font(bold=True)
                    
                    current_row = summaryStartRow + 1
                    
                    # Find account row
                    accountRow = processor.find_account_row()
                    if accountRow is None:
                        print(f"Account not found for Format2: {processor.accountName}")
                        continue
                    
                    # Initialize sums for closing balance calculation
                    sumH = 0
                    sumI = 0
                    
                    # Copy transaction data starting from accountRow + 1
                    for transactionRow in range(accountRow + 1, processor.wsTrans.max_row + 1):
                        val = processor.wsTrans.cell(row=transactionRow, column=1).value
                        if val is not None and (str(val).startswith("Total") or str(val).startswith("Closing Balance")):
                            break
                        
                        # Copy relevant transaction columns (1,2,5,7,8,9 -> 1,2,3,4,5,6)
                        processor.ws.cell(row=current_row, column=1).value = processor.wsTrans.cell(row=transactionRow, column=1).value
                        processor.ws.cell(row=current_row, column=2).value = processor.wsTrans.cell(row=transactionRow, column=2).value
                        processor.ws.cell(row=current_row, column=3).value = processor.wsTrans.cell(row=transactionRow, column=5).value
                        processor.ws.cell(row=current_row, column=4).value = processor.wsTrans.cell(row=transactionRow, column=7).value
                        processor.ws.cell(row=current_row, column=5).value = processor.wsTrans.cell(row=transactionRow, column=8).value
                        processor.ws.cell(row=current_row, column=6).value = processor.wsTrans.cell(row=transactionRow, column=9).value
                        
                        # Apply dd-mm-yyyy date format to the date column (column 1)
                        processor.ws.cell(row=current_row, column=1).number_format = "dd-mm-yyyy"
                        
                        # Update sums for closing balance calculation
                        h_val_raw = processor.wsTrans.cell(row=transactionRow, column=8).value
                        i_val_raw = processor.wsTrans.cell(row=transactionRow, column=9).value
                        sumH += safe_float(h_val_raw, f"format2 sumH row {transactionRow}")
                        sumI += safe_float(i_val_raw, f"format2 sumI row {transactionRow}")
                        
                        current_row += 1
                    
                    # Calculate closing balance with rounding to avoid floating-point precision errors
                    closingBalance = round(sumI - sumH, 2)
                    
                    # Add Closing Balance row
                    closingBalanceRow = current_row + 1
                    processor.ws.cell(row=closingBalanceRow, column=1).value = "Closing Balance"
                    if closingBalance > 0:
                        processor.ws.cell(row=closingBalanceRow, column=5).value = closingBalance
                    else:
                        processor.ws.cell(row=closingBalanceRow, column=6).value = abs(closingBalance)
                    
                    # Bold the Closing Balance row
                    for col in range(1, 7):
                        processor.ws.cell(row=closingBalanceRow, column=col).font = openpyxl.styles.Font(bold=True)
                    
                    # Update cell C8 with absolute value of closing balance (rounded)
                    processor.ws.cell(row=8, column=3).value = round(abs(closingBalance), 2)
                    
                    # Add Total row
                    totalRow = closingBalanceRow + 1
                    processor.ws.cell(row=totalRow, column=1).value = "Total"
                    processor.ws.cell(row=totalRow, column=5).value = round(sumH + (closingBalance if closingBalance > 0 else 0), 2)
                    processor.ws.cell(row=totalRow, column=6).value = round(sumI - (closingBalance if closingBalance < 0 else 0), 2)
                    
                    # Bold the Total row
                    for col in range(1, 7):
                        processor.ws.cell(row=totalRow, column=col).font = openpyxl.styles.Font(bold=True)
                    
                    # Apply summary table formatting
                    format_summary_table(processor.ws, start_row=15)
                # --- Format3 ---
                elif fmt == 'format3':
                    processor = FormatProcessorBase(target_sheet, wb["Account Transactions"])
                    
                    # Get the period end date from cell A8
                    periodEndDate = processor.ws.cell(row=8, column=1).value
                    try:
                        periodEndDate_dt = pd.to_datetime(periodEndDate, errors='raise')
                    except Exception:
                        print(f"Invalid date in A8 for Format3: {periodEndDate}")
                        continue
                    
                    # Find account and calculate balances
                    accountRow = processor.find_account_row()
                    if accountRow is None:
                        print(f"Account head not found for Format3: {processor.accountName}")
                        continue
                    
                    # Calculate opening balance (I - H)
                    openingBalance = processor.get_opening_balance(accountRow, "I_minus_H")
                    closingBalance = openingBalance
                    
                    # Loop through transactions to calculate the closing balance
                    for transactionRow in range(accountRow + 2, processor.wsTrans.max_row + 1):
                        val = processor.wsTrans.cell(row=transactionRow, column=1).value
                        if val is not None and ("Total" in str(val) or "Closing Balance" in str(val)):
                            break
                        val_i_raw = processor.wsTrans.cell(row=transactionRow, column=9).value
                        val_h_raw = processor.wsTrans.cell(row=transactionRow, column=8).value
                        val_i = safe_float(val_i_raw, f"format3 val_i row {transactionRow}")
                        val_h = safe_float(val_h_raw, f"format3 val_h row {transactionRow}")
                        closingBalance += val_i - val_h
                    
                    # Round the final closing balance to avoid floating-point precision issues
                    closingBalance = round(closingBalance, 2)
                    
                    # Setup summary table headers
                    headers = ["Date", "Particular", "£"]
                    summaryStartRow = processor.setup_summary_headers(headers, 15)
                    
                    # Fill in reconciliation data
                    processor.ws.cell(row=summaryStartRow, column=1).value = periodEndDate
                    processor.ws.cell(row=summaryStartRow, column=2).value = "Balance as per statement"
                    processor.ws.cell(row=summaryStartRow, column=3).value = ""  # Manual input placeholder
                    
                    processor.ws.cell(row=summaryStartRow + 1, column=1).value = periodEndDate
                    processor.ws.cell(row=summaryStartRow + 1, column=2).value = "Balance as per Xero"
                    processor.ws.cell(row=summaryStartRow + 1, column=3).value = closingBalance
                    
                    # Apply summary table formatting
                    format_summary_table(processor.ws, start_row=15)
                # --- Format4 ---
                elif fmt == 'format4':
                    processor = FormatProcessorBase(target_sheet, wb["Account Transactions"])
                    
                    # Find account row
                    accountRow = processor.find_account_row()
                    if accountRow is None:
                        print(f"Account not found for Format4: {processor.accountName}")
                        continue
                    
                    # Get opening balance from column I (row after account header)
                    openingBalance_raw = processor.wsTrans.cell(row=accountRow + 1, column=9).value
                    openingBalance = safe_float(openingBalance_raw, f"format4 openingBalance row {accountRow + 1}")
                    
                    # Dictionary to store monthly liability and payment data
                    monthDict = {}
                    
                    # Process transactions starting from accountRow + 2, until "Total PAYE" or "Closing Balance"
                    for transactionRow in range(accountRow + 2, processor.wsTrans.max_row + 1):
                        cell_value = processor.wsTrans.cell(row=transactionRow, column=1).value
                        if cell_value in (None, "", "Total PAYE", "Closing Balance"):
                            break
                        
                        # Extract transaction amounts
                        amountI_raw = processor.wsTrans.cell(row=transactionRow, column=9).value
                        amountH_raw = processor.wsTrans.cell(row=transactionRow, column=8).value
                        amountI = safe_float(amountI_raw, f"format4 amountI row {transactionRow}")
                        amountH = safe_float(amountH_raw, f"format4 amountH row {transactionRow}")
                        
                        # Extract transaction date and get month key
                        transactionDate = processor.wsTrans.cell(row=transactionRow, column=1).value
                        monthKey = extract_month_key(transactionDate)
                        
                        # Initialize month entry if not exists
                        if monthKey not in monthDict:
                            monthDict[monthKey] = {"liability": 0, "payment": 0}
                        
                        # Get transaction details from columns B, C, and E
                        colB = str(processor.wsTrans.cell(row=transactionRow, column=2).value or "")
                        colC = str(processor.wsTrans.cell(row=transactionRow, column=3).value or "")
                        colE = str(processor.wsTrans.cell(row=transactionRow, column=5).value or "")
                        
                        # Check if transaction is HMRC/NEST related
                        is_hmrc_nest = any(x in colC.upper() or x in colE.upper() for x in ["HMRC", "NEST"])
                        
                        # Liabilities: sum I if not HMRC/NEST
                        if not is_hmrc_nest:
                            monthDict[monthKey]["liability"] += amountI
                        
                        # Payment: sum H if HMRC/NEST
                        if is_hmrc_nest:
                            monthDict[monthKey]["payment"] += amountH
                        
                        # Subtract H from liabilities if Manual Journal in B
                        if colB.upper() == "MANUAL JOURNAL":
                            monthDict[monthKey]["liability"] -= amountH
                    
                    # Setup summary table headers
                    headers = ["Month", "Liability", "Payment", "Outstanding"]
                    summaryStartRow = processor.setup_summary_headers(headers, 15)
                    
                    # Write opening balance row
                    processor.ws.cell(row=summaryStartRow, column=1).value = "Opening Balance"
                    processor.ws.cell(row=summaryStartRow, column=2).value = openingBalance
                    processor.ws.cell(row=summaryStartRow, column=4).value = openingBalance
                    current_row = summaryStartRow + 1
                    
                    # Output each month's calculated values
                    totalOutstanding = openingBalance
                    for monthKey in monthDict:
                        liabilities = monthDict[monthKey]["liability"]
                        payment = monthDict[monthKey]["payment"]
                        outstanding = liabilities - payment
                        
                        processor.ws.cell(row=current_row, column=1).value = monthKey
                        processor.ws.cell(row=current_row, column=2).value = liabilities
                        processor.ws.cell(row=current_row, column=3).value = payment
                        processor.ws.cell(row=current_row, column=4).value = outstanding
                        
                        totalOutstanding += outstanding
                        current_row += 1
                    
                    # Add outstanding total row
                    processor.ws.cell(row=current_row, column=1).value = "Outstanding Total"
                    processor.ws.cell(row=current_row, column=4).value = totalOutstanding
                    
                    # Apply summary table formatting
                    format_summary_table(processor.ws, start_row=15)
                # --- Format5 ---
                elif fmt == 'format5':
                    processor = FormatProcessorBase(target_sheet, wb["Account Transactions"])
                    
                    # Find account row
                    accountRow = processor.find_account_row()
                    if accountRow is None:
                        print(f"Account not found for Format5: {processor.accountName}")
                        continue
                    
                    # Get opening balance from column I (row after account header)
                    openingBalance_raw = processor.wsTrans.cell(row=accountRow + 1, column=9).value
                    openingBalance = safe_float(openingBalance_raw, f"format5 openingBalance row {accountRow + 1}")
                    
                    # Dictionary to store monthly data
                    monthDict = {}
                    
                    # Process transactions starting from accountRow + 2
                    for transactionRow in range(accountRow + 2, processor.wsTrans.max_row + 1):
                        cell_value = processor.wsTrans.cell(row=transactionRow, column=1).value
                        if (cell_value in (None, "", "Closing Balance") or 
                            "Total" in str(cell_value)):
                            break
                        
                        # Extract transaction amounts
                        amountI_raw = processor.wsTrans.cell(row=transactionRow, column=9).value
                        amountH_raw = processor.wsTrans.cell(row=transactionRow, column=8).value
                        amountI = safe_float(amountI_raw, f"format5 amountI row {transactionRow}")
                        amountH = safe_float(amountH_raw, f"format5 amountH row {transactionRow}")
                        
                        # Extract transaction date and get month key
                        transactionDate = processor.wsTrans.cell(row=transactionRow, column=1).value
                        monthKey = extract_month_key(transactionDate)
                        
                        # Initialize month entry if not exists
                        if monthKey not in monthDict:
                            monthDict[monthKey] = {"totalI": 0, "totalH": 0, "payment": 0}
                        
                        # Get transaction details from columns B and C
                        colC = str(processor.wsTrans.cell(row=transactionRow, column=3).value or "")
                        colB = str(processor.wsTrans.cell(row=transactionRow, column=2).value or "")
                        
                        # Update monthly totals for liabilities and payments
                        if colC.upper() != "HMRC":
                            monthDict[monthKey]["totalI"] += amountI
                        if colB.upper() == "MANUAL JOURNAL":
                            monthDict[monthKey]["totalI"] -= amountH
                        monthDict[monthKey]["totalH"] += amountH
                        if colC.upper() == "HMRC":
                            monthDict[monthKey]["payment"] += amountH - amountI
                    
                    # Setup summary table headers
                    headers = ["Month", "Liability", "Payment", "Outstanding", "Payment Date"]
                    summaryStartRow = processor.setup_summary_headers(headers, 15)
                    
                    # Write opening balance row
                    processor.ws.cell(row=summaryStartRow, column=1).value = "Opening Balance"
                    processor.ws.cell(row=summaryStartRow, column=2).value = openingBalance
                    current_row = summaryStartRow + 1
                    
                    # Initialize running totals
                    totalLiability = openingBalance
                    totalPayment = 0
                    
                    # Output each month's calculated values
                    for monthKey in monthDict:
                        liabilities = monthDict[monthKey]["totalI"]
                        payment = monthDict[monthKey]["payment"]
                        outstanding = max(0, liabilities - payment)  # Ensure outstanding is not negative
                        
                        processor.ws.cell(row=current_row, column=1).value = monthKey
                        processor.ws.cell(row=current_row, column=2).value = liabilities
                        processor.ws.cell(row=current_row, column=3).value = payment
                        processor.ws.cell(row=current_row, column=4).value = outstanding
                        
                        totalLiability += liabilities
                        totalPayment += payment
                        current_row += 1
                    
                    # Write the Balance row
                    totalOutstanding = totalLiability - totalPayment
                    processor.ws.cell(row=current_row, column=1).value = "Balance"
                    processor.ws.cell(row=current_row, column=2).value = totalLiability
                    processor.ws.cell(row=current_row, column=3).value = totalPayment
                    processor.ws.cell(row=current_row, column=4).value = totalOutstanding
                    
                    # Update cell C8 with total outstanding
                    processor.ws.cell(row=8, column=3).value = totalOutstanding
                    processor.ws.cell(row=8, column=3).number_format = "_(* #,##0_);_(* (#,##0);_(* \"-\"??_);_(@_)"
                    
                    # Process P&L data for tax component calculations
                    wsPL = wb["P&L"] if "P&L" in wb.sheetnames else None
                    if wsPL:
                        # Initialize totals
                        netProfitBeforeTax = 0
                        depreciation = 0
                        netProfitBeforeTaxYTD = 0
                        depreciationYTD = 0
                        
                        # Process P&L sheet data
                        for row in range(1, wsPL.max_row + 1):
                            val = wsPL.cell(row=row, column=1).value
                            if val:
                                val_str = str(val)
                                col2_val = safe_float(wsPL.cell(row=row, column=2).value, f"format5 PL col2 row {row}")
                                col3_val = safe_float(wsPL.cell(row=row, column=3).value, f"format5 PL col3 row {row}")
                                
                                if "Profit after Taxation" in val_str or "Corporation Tax Expense" in val_str:
                                    netProfitBeforeTax += col2_val
                                    netProfitBeforeTaxYTD += col3_val
                                elif "Depreciation" in val_str:
                                    depreciation += col2_val
                                    depreciationYTD += col3_val
                        
                        # Calculate net profit
                        netProfit = netProfitBeforeTax + depreciation
                        netProfitYTD = netProfitBeforeTaxYTD + depreciationYTD
                        
                        # Tax rates (configurable)
                        taxRateUpTo50K = 0.25
                        taxRateAbove50K = 0.25
                        
                        # Calculate CT charges
                        def calculate_ct_charge(profit):
                            if profit < 0:
                                return 0
                            elif profit < 50000:
                                return profit * taxRateUpTo50K
                            else:
                                return profit * taxRateAbove50K
                        
                        ctChargeMonthly = calculate_ct_charge(netProfit)
                        ctChargeYTD = calculate_ct_charge(netProfitYTD)
                        
                        # Format month/year from cell A8
                        monthYear = processor.ws.cell(row=8, column=1).value
                        try:
                            monthYear_fmt = pd.to_datetime(monthYear).strftime("%b'%y")
                        except Exception:
                            monthYear_fmt = str(monthYear)
                        
                        # Create tax calculation table (columns I-K, starting row 15)
                        tax_data = [
                            ("", monthYear_fmt, "YTD"),
                            ("Net profit before tax", netProfitBeforeTax, netProfitBeforeTaxYTD),
                            ("", "", ""),
                            ("Depreciation", depreciation, depreciationYTD),
                            ("", "", ""),
                            ("Net profit", netProfit, netProfitYTD),
                            ("", "", ""),
                            ("Corporation Tax rate", taxRateUpTo50K, taxRateUpTo50K),
                            ("", "", ""),
                            ("CT charge", ctChargeMonthly, ctChargeYTD),
                            ("", "", ""),
                            ("Total CT", ctChargeMonthly, ctChargeYTD)
                        ]
                        
                        # Fill tax calculation table (columns I-K)
                        for i, (label, monthly, ytd) in enumerate(tax_data):
                            row_num = 15 + i
                            processor.ws.cell(row=row_num, column=9).value = label  # Column I
                            processor.ws.cell(row=row_num, column=10).value = monthly  # Column J
                            processor.ws.cell(row=row_num, column=11).value = ytd  # Column K
                            
                            # Bold the header row
                            if i == 0:
                                for col in range(9, 12):  # Columns I-K
                                    processor.ws.cell(row=row_num, column=col).font = openpyxl.styles.Font(bold=True)
                            
                            # Apply accounting format to numeric values
                            if isinstance(monthly, (int, float)) and monthly != 0:
                                processor.ws.cell(row=row_num, column=10).number_format = "#,##0.00_);(#,##0.00);-_);_(@_)"
                            if isinstance(ytd, (int, float)) and ytd != 0:
                                processor.ws.cell(row=row_num, column=11).number_format = "#,##0.00_);(#,##0.00);-_);_(@_)"
                    
                    # Apply summary table formatting (for columns A-E only)
                    # Custom formatting to avoid interfering with tax table in columns G-I
                    from openpyxl.styles import Font, Alignment, Border, Side
                    
                    start_row = 15
                    max_row = processor.ws.max_row
                    last_row = start_row
                    last_col = 5  # Only format columns A-E (1-5)
                    
                    # Find the actual last row with data in columns A-E
                    for r in range(start_row, max_row + 1):
                        row_has_data = False
                        for c in range(1, 6):  # Columns A-E
                            val = processor.ws.cell(row=r, column=c).value
                            if val is not None and str(val).strip() != "":
                                row_has_data = True
                        if row_has_data:
                            last_row = r
                    
                    # Apply formatting only to columns A-E
                    if last_row >= start_row:
                        font = Font(name="Gill Sans MT", size=12)
                        bold_font = Font(name="Gill Sans MT", size=12, bold=True)
                        border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
                        
                        for r in range(start_row, last_row + 1):
                            for c in range(1, last_col + 1):  # Only columns A-E
                                cell = processor.ws.cell(row=r, column=c)
                                val = cell.value
                                if r == start_row:
                                    cell.font = bold_font
                                    cell.alignment = Alignment(horizontal="center", vertical="bottom")
                                else:
                                    cell.font = font
                                    if isinstance(val, (int, float)):
                                        cell.alignment = Alignment(horizontal="right", vertical="bottom")
                                        if val is not None:
                                            if val < 0:
                                                cell.number_format = "#,##0;(#,##0)"
                                            elif val == 0:
                                                cell.number_format = "#,##0;-#,##0;-"
                                            else:
                                                cell.number_format = "#,##0"
                                    else:
                                        cell.alignment = Alignment(horizontal="left", vertical="bottom")
                                cell.border = border
                        
                        # Auto-adjust column widths for columns A-E only
                        for c in range(1, last_col + 1):
                            maxlen = 0
                            for r in range(start_row, last_row + 1):
                                v = processor.ws.cell(row=r, column=c).value
                                vstr = str(v) if v is not None else ""
                                if len(vstr) > maxlen:
                                    maxlen = len(vstr)
                            processor.ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = max(10, min(maxlen + 2, 40))
                    
                    # Apply tax component formatting (for columns I-K)
                    apply_tax_component_formatting(processor.ws)
                # --- Format6 ---
                elif fmt == 'format6':
                    processor = FormatProcessorBase(target_sheet, wb["Account Transactions"])
                    
                    # Find account row
                    accountRow = processor.find_account_row()
                    if accountRow is None:
                        print(f"Account not found for Format6: {processor.accountName}")
                        continue
                    
                    # Get opening balance from column I (row after account header)
                    openingBalance_raw = processor.wsTrans.cell(row=accountRow + 1, column=9).value
                    openingBalance = safe_float(openingBalance_raw, f"format6 openingBalance row {accountRow + 1}")
                    
                    # Dictionary to store monthly data and running totals
                    monthDict = {}
                    
                    # Process transactions starting from accountRow + 2
                    for transactionRow in range(accountRow + 2, processor.wsTrans.max_row + 1):
                        cell_value = processor.wsTrans.cell(row=transactionRow, column=1).value
                        if (cell_value in (None, "", "Closing Balance") or 
                            "Total" in str(cell_value)):
                            break
                        
                        # Extract transaction amounts
                        amountI_raw = processor.wsTrans.cell(row=transactionRow, column=9).value
                        amountH_raw = processor.wsTrans.cell(row=transactionRow, column=8).value
                        amountI = safe_float(amountI_raw, f"format6 amountI row {transactionRow}")
                        amountH = safe_float(amountH_raw, f"format6 amountH row {transactionRow}")
                        
                        # Extract transaction date and get month key
                        transactionDate = processor.wsTrans.cell(row=transactionRow, column=1).value
                        monthKey = extract_month_key(transactionDate)
                        
                        # Initialize month entry if not exists
                        if monthKey not in monthDict:
                            monthDict[monthKey] = {"totalI": 0, "totalH": 0}
                        
                        # Update monthly totals for liabilities based on Column C (excluding "HMRC")
                        colC = str(processor.wsTrans.cell(row=transactionRow, column=3).value or "")
                        if colC.upper() != "HMRC":
                            monthDict[monthKey]["totalI"] += amountI
                        monthDict[monthKey]["totalH"] += amountH
                    
                    # Setup summary table headers
                    headers = ["Description", "Liability", "Payment", "Difference"]
                    summaryStartRow = processor.setup_summary_headers(headers, 15)
                    
                    # Write opening balance row
                    processor.ws.cell(row=summaryStartRow, column=1).value = "Opening Balance"
                    processor.ws.cell(row=summaryStartRow, column=2).value = openingBalance
                    processor.ws.cell(row=summaryStartRow, column=4).value = openingBalance
                    current_row = summaryStartRow + 1
                    
                    # Initialize running totals
                    totalLiability = openingBalance
                    totalPayment = 0
                    
                    # Output each month's calculated values
                    for monthKey in monthDict:
                        liabilities = monthDict[monthKey]["totalI"]
                        payment = monthDict[monthKey]["totalH"]
                        difference = liabilities - payment
                        
                        processor.ws.cell(row=current_row, column=1).value = monthKey
                        processor.ws.cell(row=current_row, column=2).value = liabilities
                        processor.ws.cell(row=current_row, column=3).value = payment
                        processor.ws.cell(row=current_row, column=4).value = difference
                        
                        totalLiability += liabilities
                        totalPayment += payment
                        current_row += 1
                    
                    # Write the Balance row
                    totalOutstanding = totalLiability - totalPayment
                    processor.ws.cell(row=current_row, column=1).value = "Balance"
                    processor.ws.cell(row=current_row, column=2).value = totalLiability
                    processor.ws.cell(row=current_row, column=3).value = totalPayment
                    processor.ws.cell(row=current_row, column=4).value = totalOutstanding
                    
                    # Apply summary table formatting
                    format_summary_table(processor.ws, start_row=15)
                # --- Format7 ---
                elif fmt == 'format7':
                    processor = FormatProcessorBase(target_sheet, wb["Account Transactions"])
                    
                    # Find account row
                    accountRow = processor.find_account_row()
                    if accountRow is None:
                        print(f"Account not found for Format7: {processor.accountName}")
                        continue
                    
                    # Get opening balance from column I (row after account header)
                    openingBalance_raw = processor.wsTrans.cell(row=accountRow + 1, column=9).value
                    openingBalance = safe_float(openingBalance_raw, f"format7 openingBalance row {accountRow + 1}")
                    
                    # Dictionary to store monthly liability and payment data
                    monthDict = {}
                    
                    # Process transactions starting from accountRow + 2
                    for transactionRow in range(accountRow + 2, processor.wsTrans.max_row + 1):
                        cell_value = processor.wsTrans.cell(row=transactionRow, column=1).value
                        if (cell_value in (None, "", "Closing Balance") or 
                            str(cell_value).startswith("Total")):
                            break
                        
                        # Extract transaction amounts
                        amountI_raw = processor.wsTrans.cell(row=transactionRow, column=9).value
                        amountH_raw = processor.wsTrans.cell(row=transactionRow, column=8).value
                        amountI = safe_float(amountI_raw, f"format7 amountI row {transactionRow}")
                        amountH = safe_float(amountH_raw, f"format7 amountH row {transactionRow}")
                        
                        # Extract transaction date and get month key
                        transactionDate = processor.wsTrans.cell(row=transactionRow, column=1).value
                        monthKey = extract_month_key(transactionDate)
                        
                        # Initialize month entry if not exists
                        if monthKey not in monthDict:
                            monthDict[monthKey] = {"liability": 0, "payment": 0}
                        
                        # Get transaction type from column B
                        colB = str(processor.wsTrans.cell(row=transactionRow, column=2).value or "")
                        
                        # Categorize transaction by type
                        if colB == "Manual Journal":
                            monthDict[monthKey]["liability"] += amountI - amountH
                        elif colB == "Spend Money":
                            monthDict[monthKey]["payment"] += amountH
                    
                    # Setup summary table headers
                    headers = ["Month", "Liability", "Payment", "Outstanding"]
                    summaryStartRow = processor.setup_summary_headers(headers, 15)
                    
                    # Write opening balance row
                    processor.ws.cell(row=summaryStartRow, column=1).value = "Opening Balance"
                    processor.ws.cell(row=summaryStartRow, column=2).value = openingBalance
                    processor.ws.cell(row=summaryStartRow, column=4).value = openingBalance
                    current_row = summaryStartRow + 1
                    
                    # Output each month's calculated values
                    totalOutstanding = openingBalance
                    for monthKey in monthDict:
                        liabilities = monthDict[monthKey]["liability"]
                        payment = monthDict[monthKey]["payment"]
                        outstanding = liabilities - payment
                        
                        processor.ws.cell(row=current_row, column=1).value = monthKey
                        processor.ws.cell(row=current_row, column=2).value = liabilities
                        processor.ws.cell(row=current_row, column=3).value = payment
                        
                        # Display Outstanding only if nonzero, else leave blank
                        if outstanding != 0:
                            processor.ws.cell(row=current_row, column=4).value = outstanding
                        else:
                            processor.ws.cell(row=current_row, column=4).value = None
                        
                        totalOutstanding += outstanding
                        current_row += 1
                    
                    # Add outstanding total row
                    processor.ws.cell(row=current_row, column=1).value = "Outstanding Total"
                    processor.ws.cell(row=current_row, column=4).value = totalOutstanding
                    
                    # Apply summary table formatting
                    format_summary_table(processor.ws, start_row=15)
                # --- Format8 ---
                elif fmt == 'format8':
                    processor = FormatProcessorBase(target_sheet, wb["Account Transactions"])
                    
                    # Find account row
                    accountRow = processor.find_account_row()
                    if accountRow is None:
                        print(f"Account not found for Format8: {processor.accountName}")
                        continue
                    
                    # Get opening balance from column H (row after account header)
                    openingBalance_raw = processor.wsTrans.cell(row=accountRow + 1, column=8).value
                    openingBalance = safe_float(openingBalance_raw, f"format8 openingBalance row {accountRow + 1}")
                    closingBalance = openingBalance
                    
                    # Dictionary to store monthly breakdowns
                    monthDict = {}
                    
                    # Process transactions starting from accountRow + 2
                    for transactionRow in range(accountRow + 2, processor.wsTrans.max_row + 1):
                        cell_value = processor.wsTrans.cell(row=transactionRow, column=1).value
                        if cell_value in (None, "", "Total PAYE", "Closing Balance"):
                            break
                        
                        # Extract transaction amounts
                        amountI_raw = processor.wsTrans.cell(row=transactionRow, column=9).value
                        amountH_raw = processor.wsTrans.cell(row=transactionRow, column=8).value
                        amountI = safe_float(amountI_raw, f"format8 amountI row {transactionRow}")
                        amountH = safe_float(amountH_raw, f"format8 amountH row {transactionRow}")
                        
                        # Extract transaction date and get month key
                        transactionDate = processor.wsTrans.cell(row=transactionRow, column=1).value
                        monthKey = extract_month_key(transactionDate)
                        
                        # Initialize month entry if not exists
                        if monthKey not in monthDict:
                            monthDict[monthKey] = {"receipts": 0, "payments": 0, "pdqDeposits": 0}
                        
                        # Get transaction type from column B
                        colB = str(processor.wsTrans.cell(row=transactionRow, column=2).value or "").upper()
                        
                        # Categorize transaction by type
                        if colB == "RECEIVE MONEY":
                            monthDict[monthKey]["receipts"] += amountH
                        elif colB in ("SPEND MONEY", "PAYABLE PAYMENT", "PAYABLE OVERPAYMENT"):
                            monthDict[monthKey]["payments"] += amountI
                        elif colB == "BANK TRANSFER":
                            monthDict[monthKey]["pdqDeposits"] += (amountI - amountH)
                    
                    # Setup summary table headers
                    headers = ["Month", "Op Bal", "Receipts", "Payments", "PDQ / Deposits", "Clo Bal"]
                    summaryStartRow = processor.setup_summary_headers(headers, 15)
                    
                    # Write opening balance row
                    processor.ws.cell(row=summaryStartRow, column=1).value = "Opening Balance"
                    processor.ws.cell(row=summaryStartRow, column=2).value = openingBalance
                    current_row = summaryStartRow + 1
                    
                    # Output each month's calculated values
                    for monthKey in monthDict:
                        receipts = monthDict[monthKey]["receipts"]
                        payments = monthDict[monthKey]["payments"]
                        pdqDeposits = monthDict[monthKey]["pdqDeposits"]
                        closingBalance = closingBalance + receipts - payments - pdqDeposits
                        
                        processor.ws.cell(row=current_row, column=1).value = monthKey
                        processor.ws.cell(row=current_row, column=2).value = openingBalance
                        processor.ws.cell(row=current_row, column=3).value = receipts
                        processor.ws.cell(row=current_row, column=4).value = payments
                        processor.ws.cell(row=current_row, column=5).value = pdqDeposits
                        processor.ws.cell(row=current_row, column=6).value = closingBalance
                        
                        # Update opening balance for next month
                        openingBalance = closingBalance
                        current_row += 1
                    
                    # Apply summary table formatting
                    format_summary_table(processor.ws, start_row=15)
                # --- Format9 ---
                elif fmt == 'format9':
                    processor = FormatProcessorBase(target_sheet, wb["Account Transactions"])
                    
                    # Get the period end date from cell A8
                    periodEndDate = processor.ws.cell(row=8, column=1).value
                    try:
                        periodEndDate_dt = pd.to_datetime(periodEndDate, errors='raise')
                    except Exception:
                        print(f"Invalid date in A8 for Format9: {periodEndDate}")
                        continue
                    
                    # Find account and calculate balances
                    accountRow = processor.find_account_row()
                    if accountRow is None:
                        print(f"Account head not found for Format9: {processor.accountName}")
                        continue
                    
                    # Calculate opening balance (H - I) 
                    openingBalance = processor.get_opening_balance(accountRow, "H_minus_I")
                    closingBalance = openingBalance
                    
                    # Loop through transactions to calculate the closing balance
                    for transactionRow in range(accountRow + 2, processor.wsTrans.max_row + 1):
                        val = processor.wsTrans.cell(row=transactionRow, column=1).value
                        if val is not None and ("Total" in str(val) or "Closing Balance" in str(val)):
                            break
                        creditVal_raw = processor.wsTrans.cell(row=transactionRow, column=8).value
                        debitVal_raw = processor.wsTrans.cell(row=transactionRow, column=9).value
                        creditVal = safe_float(creditVal_raw, f"format9 creditVal row {transactionRow}")
                        debitVal = safe_float(debitVal_raw, f"format9 debitVal row {transactionRow}")
                        closingBalance = closingBalance + (creditVal - debitVal)
                    
                    # Round the final closing balance to avoid floating-point precision issues
                    closingBalance = round(closingBalance, 2)
                    
                    # Setup summary table
                    headers = ["Date", "Details", "Amount £"]
                    summaryStartRow = processor.setup_summary_headers(headers, 15)
                    
                    # Fill in summary data
                    processor.ws.cell(row=summaryStartRow, column=1).value = periodEndDate
                    processor.ws.cell(row=summaryStartRow, column=1).number_format = "dd-mm-yyyy"
                    processor.ws.cell(row=summaryStartRow, column=2).value = "Balance as per statement"
                    processor.ws.cell(row=summaryStartRow, column=3).value = ""
                    summaryStartRow += 2  # Skip blank row
                    
                    processor.ws.cell(row=summaryStartRow, column=1).value = periodEndDate
                    processor.ws.cell(row=summaryStartRow, column=1).number_format = "dd-mm-yyyy"
                    processor.ws.cell(row=summaryStartRow, column=2).value = "Balance per Control account"
                    processor.ws.cell(row=summaryStartRow, column=3).value = closingBalance
                    processor.ws.cell(row=summaryStartRow, column=3).number_format = "#,##0.00"
                    
                    # Apply formatting
                    format_summary_table(processor.ws, start_row=15)
                # --- Format10 ---
                elif fmt == 'format10':
                    processor = FormatProcessorBase(target_sheet, wb["Account Transactions"])
                    
                    # Get the period end date from cell A8
                    periodEndDate = processor.ws.cell(row=8, column=1).value
                    try:
                        periodEndDate_dt = pd.to_datetime(periodEndDate, errors='raise')
                    except Exception:
                        print(f"Invalid date in A8 for Format10: {periodEndDate}")
                        continue
                    
                    # Find account and calculate balances
                    accountRow = processor.find_account_row()
                    if accountRow is None:
                        print(f"Account head not found for Format10: {processor.accountName}")
                        continue
                    
                    # Calculate opening balance (I - H)
                    openingBalance = processor.get_opening_balance(accountRow, "I_minus_H")
                    closingBalance = openingBalance
                    
                    # Loop through transactions to calculate the closing balance
                    for transactionRow in range(accountRow + 2, processor.wsTrans.max_row + 1):
                        val = processor.wsTrans.cell(row=transactionRow, column=1).value
                        if val is not None and ("Total" in str(val) or "Closing Balance" in str(val)):
                            break
                        debitVal_raw = processor.wsTrans.cell(row=transactionRow, column=9).value
                        creditVal_raw = processor.wsTrans.cell(row=transactionRow, column=8).value
                        debitVal = safe_float(debitVal_raw, f"format10 debitVal row {transactionRow}")
                        creditVal = safe_float(creditVal_raw, f"format10 creditVal row {transactionRow}")
                        closingBalance = closingBalance + (debitVal - creditVal)
                    
                    # Round the final closing balance to avoid floating-point precision issues
                    closingBalance = round(closingBalance, 2)
                    
                    # Setup reconciliation header
                    processor.ws.cell(row=13, column=1).value = "Reconciliation"
                    processor.ws.cell(row=13, column=1).font = openpyxl.styles.Font(bold=True, size=14)
                    
                    # Setup summary table headers
                    headers = ["Date", "£", "Particular"]
                    summaryStartRow = processor.setup_summary_headers(headers, 15)
                    
                    # Fill in reconciliation data
                    processor.ws.cell(row=summaryStartRow, column=1).value = periodEndDate
                    processor.ws.cell(row=summaryStartRow, column=1).number_format = "dd-mm-yyyy"
                    processor.ws.cell(row=summaryStartRow, column=2).value = ""  # Manual input placeholder
                    processor.ws.cell(row=summaryStartRow, column=3).value = "Balance as per "
                    
                    processor.ws.cell(row=summaryStartRow + 1, column=1).value = periodEndDate
                    processor.ws.cell(row=summaryStartRow + 1, column=1).number_format = "dd-mm-yyyy"
                    processor.ws.cell(row=summaryStartRow + 1, column=2).value = closingBalance
                    processor.ws.cell(row=summaryStartRow + 1, column=3).value = "Balance as per "
                    
                    processor.ws.cell(row=summaryStartRow + 2, column=1).value = periodEndDate
                    processor.ws.cell(row=summaryStartRow + 2, column=1).number_format = "dd-mm-yyyy"
                    processor.ws.cell(row=summaryStartRow + 2, column=2).value = f"=B{summaryStartRow}-B{summaryStartRow + 1}"
                    processor.ws.cell(row=summaryStartRow + 2, column=3).value = "Difference"
                    
                    # Apply accounting format to amount column
                    for r in range(summaryStartRow, summaryStartRow + 3):
                        processor.ws.cell(row=r, column=2).number_format = "#,##0.00"
                    
                    # Apply formatting
                    format_summary_table(processor.ws, start_row=15)



    # --- Ensure FAR tables, depreciation rates, and months are built from the current workbook ---
    if 'FAR' in wb.sheetnames:
        ws_far = wb['FAR']
        # Remove gridlines from FAR sheet
        ws_far.sheet_view.showGridLines = False
        # Extract FAR data from the current workbook with fresh stream
        df_raw = pd.read_excel(create_fresh_stream(original_content), sheet_name='FAR', header=None, engine='openpyxl')
        static_headers = [
            'Purchase Date',
            'Details',
            'Cost',
            'Addition',
            'Total Cost',
            'Depreciation Rate',
            'Accumulated Depreciation'
        ]
        # Generate fiscal year month columns as strings like "Dep    Aug-24"
        months = []
        cur = fy_start.replace(day=1)
        while cur <= fy_end:
            months.append(f"Dep {cur.strftime('%b-%y')}")
            cur += pd.DateOffset(months=1)
        tables = []
        far_table_names = []
        far_table_deprates = {}
        i = 5  # Row 6 in Excel (0-based index)
        while i < len(df_raw):
            table_name = str(df_raw.iloc[i, 0])
            if table_name == 'nan' or not table_name.strip():
                i += 1
                continue
            far_table_names.append(table_name.strip())
            depreciation_row = i + 1
            dep_rate_val = None
            if depreciation_row < len(df_raw):
                dep_line = str(df_raw.iloc[depreciation_row, 0])
               
                match = re.search(r"Depreciation rate\s*:\s*([\d.]+)%", dep_line, re.IGNORECASE)
                if match:
                    dep_rate_val = float(match.group(1))
            if dep_rate_val is not None:
                far_table_deprates[table_name.strip()] = dep_rate_val
            header_row = i + 2
            data_start = i + 3
            data_end = data_start
            while data_end < len(df_raw):
                if str(df_raw.iloc[data_end, 1]).strip().lower() == 'total':
                    break
                data_end += 1
            table_data = df_raw.iloc[data_start:data_end, :len(static_headers)]
            table_data.columns = static_headers
            table_data = table_data.reset_index(drop=True)
            for m in months:
                table_data[m] = 0
            tables.append((table_name, table_data))
            i = data_end + 1
            while i < len(df_raw) and (str(df_raw.iloc[i, 0]) == 'nan' or not str(df_raw.iloc[i, 0]).strip()):
                i += 1

        # --- Merge new transactions and recalculate FAR tables before writing ---
        # Read Account Transactions sheet from the same Excel as FAR
        if 'Account Transactions' in wb.sheetnames:
            ws_trans = wb['Account Transactions']
            # Read as DataFrame for easier processing with fresh stream
            df_trans_raw = pd.read_excel(create_fresh_stream(original_content), sheet_name='Account Transactions', header=None, engine='openpyxl')
            header_row_idx = 4  # Row 5 in Excel (0-based)
            headers = list(df_trans_raw.iloc[header_row_idx].fillna('').astype(str))
            header_map = {h.strip().lower(): i for i, h in enumerate(headers)}
            transactions = []
            i = header_row_idx + 1
            while i < len(df_trans_raw):
                account_name = str(df_trans_raw.iloc[i, 0])
                if account_name == 'nan' or not account_name.strip() or account_name.strip().lower() == 'total':
                    i += 1
                    continue
                if account_name not in far_table_names:
                    i += 1
                    continue
                data_start = i + 1
                data_end = data_start
                while data_end < len(df_trans_raw):
                    cell_val = str(df_trans_raw.iloc[data_end, 0]).strip().lower()
                    if cell_val.startswith('total'):
                        break
                    data_end += 1
                block = df_trans_raw.iloc[data_start:data_end, :].copy()
                block = block[~block.iloc[:,0].astype(str).str.strip().str.lower().isin(['opening balance','closing balance'])]
                if not block.empty:
                    mapped = pd.DataFrame()
                    mapped['Purchase Date'] = block.iloc[:, header_map['purchase date']] if 'purchase date' in header_map else block.iloc[:, 0]
                    mapped['Details'] = block.iloc[:, header_map['details']] if 'details' in header_map else block.iloc[:, 2]
                    mapped['Cost'] = block.iloc[:, 7] if block.shape[1] > 7 else 0
                    mapped['Asset Type'] = account_name.strip()
                    transactions.append(mapped)
                i = data_end + 1
            if transactions:
                df_trans = pd.concat(transactions, ignore_index=True)
            else:
                df_trans = pd.DataFrame()
        else:
            df_trans = pd.DataFrame()

        updated_tables = []
        for table_name, table_df in tables:
            updated_table = table_df.copy()
            # Note: Do NOT filter out 'opening balance' entries from FAR tables
            # as these are legitimate asset records, unlike Account Transactions
            if 'Total Depreciation' not in updated_table.columns:
                updated_table['Total Depreciation'] = 0
            if 'WDV' not in updated_table.columns:
                updated_table['WDV'] = 0
            if not df_trans.empty and 'Asset Type' in df_trans.columns:
                matching_rows = df_trans[df_trans['Asset Type'] == table_name].reset_index(drop=True)
                if not matching_rows.empty:
                    new_rows = pd.DataFrame()
                    new_rows['Purchase Date'] = matching_rows['Purchase Date']
                    new_rows['Details'] = matching_rows['Details']
                    new_rows['Cost'] = matching_rows['Cost']
                    new_rows['Addition'] = 0
                    new_rows['Total Cost'] = 0
                    new_rows['Depreciation Rate'] = far_table_deprates.get(table_name.strip(), 0)
                    new_rows['Accumulated Depreciation'] = 0
                    for m in months:
                        new_rows[m] = 0
                    if 'Details' in new_rows.columns:
                        new_rows = new_rows[~new_rows['Details'].astype(str).str.strip().str.lower().isin(['opening balance','closing balance'])]
                    if not updated_table.empty:
                        merge_cols = ['Purchase Date', 'Details', 'Cost']
                        def safe_str(x):
                            if pd.isnull(x):
                                return ''
                            if isinstance(x, pd.Timestamp) or isinstance(x, datetime):
                                return x.strftime('%Y-%m-%d')
                            try:
                                dt = pd.to_datetime(x, errors='coerce')
                                if pd.notnull(dt):
                                    return dt.strftime('%Y-%m-%d')
                            except Exception:
                                pass
                            return str(x)
                        updated_table_key = updated_table[merge_cols].applymap(safe_str).agg('|'.join, axis=1)
                        new_rows_key = new_rows[merge_cols].applymap(safe_str).agg('|'.join, axis=1)
                        new_rows = new_rows[~new_rows_key.isin(updated_table_key)]
                    for idx, row in new_rows.iterrows():
                        try:
                            pdate = pd.to_datetime(row['Purchase Date'], errors='coerce')
                        except Exception:
                            pdate = None
                        purchase_amt = pd.to_numeric(row['Cost'], errors='coerce') if pd.notnull(row['Cost']) else 0
                        if pd.notnull(pdate) and fy_start <= pdate <= fy_end:
                            new_rows.at[idx, 'Addition'] = purchase_amt
                            new_rows.at[idx, 'Cost'] = 0
                        else:
                            new_rows.at[idx, 'Addition'] = 0
                            new_rows.at[idx, 'Cost'] = purchase_amt
                    new_rows['Total Cost'] = pd.to_numeric(new_rows['Cost'], errors='coerce').fillna(0) + pd.to_numeric(new_rows['Addition'], errors='coerce').fillna(0)
                    for idx, row in new_rows.iterrows():
                        try:
                            pdate = pd.to_datetime(row['Purchase Date'], errors='coerce')
                        except Exception:
                            pdate = None
                        total_cost = pd.to_numeric(row['Total Cost'], errors='coerce')
                        dep_rate = pd.to_numeric(row['Depreciation Rate'], errors='coerce')
                        last_fy_month = fy_start - pd.DateOffset(months=1)
                        months_in_use = 0
                        if pd.notnull(pdate) and pdate <= last_fy_month:
                            months_in_use = (last_fy_month.year - pdate.year) * 12 + (last_fy_month.month - pdate.month) + 1
                            if months_in_use < 0:
                                months_in_use = 0
                        monthly_dep = (total_cost * dep_rate / 100) / 12 if total_cost and dep_rate else 0
                        acc_dep = monthly_dep * months_in_use
                        new_rows.at[idx, 'Accumulated Depreciation'] = acc_dep
                    for idx, row in new_rows.iterrows():
                        try:
                            pdate = pd.to_datetime(row['Purchase Date'], errors='coerce')
                        except Exception:
                            pdate = None
                        total_cost = pd.to_numeric(row['Total Cost'], errors='coerce')
                        dep_rate = pd.to_numeric(row['Depreciation Rate'], errors='coerce')
                        monthly_dep = (total_cost * dep_rate / 100) / 12 if total_cost and dep_rate else 0
                        for m in months:
                            try:
                                m_str = m.replace('Dep', '').strip()
                                m_dt = pd.to_datetime(m_str, format='%b-%y')
                            except Exception:
                                m_dt = None
                            if pd.notnull(pdate) and pd.notnull(m_dt):
                                if (pdate <= m_dt + pd.offsets.MonthEnd(0)) and (m_dt <= mgmt_acct_month):
                                    new_rows.at[idx, m] = monthly_dep
                                elif m_dt > mgmt_acct_month:
                                    new_rows.at[idx, m] = ''
                                else:
                                    new_rows.at[idx, m] = 0
                            else:
                                new_rows.at[idx, m] = ''
                    dep_month_cols = months
                    acc_dep_col_name = 'Accumulated Depreciation'
                    total_cost_col_name = 'Total Cost'
                    new_rows['Total Depreciation'] = pd.to_numeric(new_rows[acc_dep_col_name], errors='coerce').fillna(0) + new_rows[dep_month_cols].apply(pd.to_numeric, errors='coerce').fillna(0).sum(axis=1)
                    new_rows['WDV'] = pd.to_numeric(new_rows[total_cost_col_name], errors='coerce').fillna(0) - new_rows['Total Depreciation']
                    if 'Total Depreciation' not in updated_table.columns:
                        updated_table['Total Depreciation'] = 0
                    if 'WDV' not in updated_table.columns:
                        updated_table['WDV'] = 0
                    full_cols = static_headers + months + ['Total Depreciation', 'WDV']
                    updated_table = updated_table.reindex(columns=full_cols, fill_value=0)
                    new_rows = new_rows.reindex(columns=full_cols, fill_value=0)
                    if not new_rows.empty:
                        print(f"Appending {len(new_rows)} new transaction row(s) to table '{table_name}'")
                    combined_table = pd.concat([updated_table, new_rows], ignore_index=True)
                    updated_table = combined_table
            dep_rate_val = far_table_deprates.get(table_name.strip(), 0)
            for col in ['Addition', 'Total Cost', 'Depreciation Rate', 'Accumulated Depreciation'] + months:
                if col not in updated_table.columns:
                    updated_table[col] = 0
            updated_table['Depreciation Rate'] = dep_rate_val
            for idx, row in updated_table.iterrows():
                try:
                    pdate = pd.to_datetime(row['Purchase Date'], errors='coerce')
                except Exception:
                    pdate = None
                cost_val = pd.to_numeric(row['Cost'], errors='coerce') if pd.notnull(row['Cost']) else 0
                addition_val = pd.to_numeric(row['Addition'], errors='coerce') if pd.notnull(row['Addition']) else 0
                orig_amt = cost_val + addition_val
                if pd.notnull(pdate) and fy_start <= pdate <= fy_end:
                    updated_table.at[idx, 'Addition'] = orig_amt
                    updated_table.at[idx, 'Cost'] = 0
                else:
                    updated_table.at[idx, 'Addition'] = 0
                    updated_table.at[idx, 'Cost'] = orig_amt
            updated_table['Total Cost'] = pd.to_numeric(updated_table['Cost'], errors='coerce').fillna(0) + pd.to_numeric(updated_table['Addition'], errors='coerce').fillna(0)
            for idx, row in updated_table.iterrows():
                try:
                    pdate = pd.to_datetime(row['Purchase Date'], errors='coerce')
                except Exception:
                    pdate = None
                total_cost = pd.to_numeric(row['Total Cost'], errors='coerce')
                dep_rate = pd.to_numeric(row['Depreciation Rate'], errors='coerce')
                last_fy_month = fy_start - pd.DateOffset(months=1)
                months_in_use = 0
                if pd.notnull(pdate) and pdate <= last_fy_month:
                    months_since_purchase = (last_fy_month.year - pdate.year) * 12 + (last_fy_month.month - pdate.month) + 1
                    if months_since_purchase < 0:
                        months_since_purchase = 0
                    monthly_dep = (total_cost * dep_rate / 100) / 12 if total_cost and dep_rate else 0
                    months_to_fully_depreciate = int(total_cost // monthly_dep) if monthly_dep > 0 else 0
                    months_in_use = min(months_since_purchase, months_to_fully_depreciate)
                    acc_dep = monthly_dep * months_in_use
                    if months_since_purchase >= months_to_fully_depreciate:
                        acc_dep = total_cost
                    updated_table.at[idx, 'Accumulated Depreciation'] = acc_dep
                else:
                    monthly_dep = (total_cost * dep_rate / 100) / 12 if total_cost and dep_rate else 0
                    updated_table.at[idx, 'Accumulated Depreciation'] = 0
            for idx2, row2 in updated_table.iterrows():
                try:
                    pdate2 = pd.to_datetime(row2['Purchase Date'], errors='coerce')
                except Exception:
                    pdate2 = None
                total_cost2 = pd.to_numeric(row2['Total Cost'], errors='coerce')
                dep_rate2 = pd.to_numeric(row2['Depreciation Rate'], errors='coerce')
                monthly_dep2 = (total_cost2 * dep_rate2 / 100) / 12 if total_cost2 and dep_rate2 else 0
                for m in months:
                    updated_table.at[idx2, m] = 0
                total_dep_so_far2 = pd.to_numeric(updated_table.at[idx2, 'Accumulated Depreciation'], errors='coerce') if 'Accumulated Depreciation' in updated_table.columns else 0
                for m in months:
                    try:
                        m_str = m.replace('Dep', '').strip()
                        m_dt = pd.to_datetime(m_str, format='%b-%y')
                    except Exception:
                        m_dt = None
                    if pd.notnull(pdate2) and pd.notnull(m_dt):
                        if (pdate2 <= m_dt + pd.offsets.MonthEnd(0)) and (m_dt <= mgmt_acct_month):
                            if total_dep_so_far2 < total_cost2:
                                dep_this_month2 = min(monthly_dep2, total_cost2 - total_dep_so_far2)
                                updated_table.at[idx2, m] = dep_this_month2
                                total_dep_so_far2 += dep_this_month2
                            else:
                                updated_table.at[idx2, m] = 0
                        elif m_dt > mgmt_acct_month:
                            updated_table.at[idx2, m] = ''
                        else:
                            updated_table.at[idx2, m] = 0
                    else:
                        updated_table.at[idx2, m] = ''
            dep_month_cols = months
            acc_dep_col_name = 'Accumulated Depreciation'
            total_cost_col_name = 'Total Cost'
            updated_table['Total Depreciation'] = pd.to_numeric(updated_table[acc_dep_col_name], errors='coerce').fillna(0) + updated_table[dep_month_cols].apply(pd.to_numeric, errors='coerce').fillna(0).sum(axis=1)
            updated_table['WDV'] = pd.to_numeric(updated_table[total_cost_col_name], errors='coerce').fillna(0) - updated_table['Total Depreciation']
            final_cols = static_headers + months + ['Total Depreciation', 'WDV']
            updated_table = updated_table.reindex(columns=final_cols, fill_value=0)
            updated_tables.append((table_name, updated_table.copy()))

        # Now write all updated_tables to the consolidated output (as before)
        from openpyxl.styles.borders import Border, Side
        from openpyxl.styles import Font, Alignment
        for row in ws_far.iter_rows(min_row=6, max_row=ws_far.max_row):
            for cell in row:
                cell.value = None
                cell.border = Border()
        write_row = 6
        if updated_tables:
            for table_name, combined_df in updated_tables:
                combined_df = combined_df.dropna(how='all').reset_index(drop=True)
                combined_df = combined_df.applymap(lambda x: x.item() if hasattr(x, 'item') else x)
                header_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                uniform_font = Font(name='Gill Sans MT', size=12, bold=False)
                bold_font = Font(name='Gill Sans MT', size=12, bold=True)
                center_align = Alignment(horizontal='center', vertical='center')
                right_align = Alignment(horizontal='right', vertical='center')
                left_align = Alignment(horizontal='left', vertical='center')
                date_align = Alignment(horizontal='center', vertical='center')
                table_name_cell = ws_far.cell(row=write_row, column=1, value=table_name)
                table_name_cell.font = bold_font
                for col in range(2, len(combined_df.columns)+1):
                    ws_far.cell(row=write_row, column=col).font = uniform_font
                    ws_far.cell(row=write_row, column=col).border = Border()
                dep_rate_val = None
                for tname, rate in far_table_deprates.items():
                    if tname == table_name.strip():
                        dep_rate_val = rate
                        break
                dep_line = f"Depreciation rate: {dep_rate_val:.0f}%" if dep_rate_val is not None else "Depreciation rate: "
                dep_line_cell = ws_far.cell(row=write_row+1, column=1, value=dep_line)
                dep_line_cell.font = uniform_font
                dep_line_cell.alignment = left_align
                for col in range(2, len(combined_df.columns)+1):
                    ws_far.cell(row=write_row+1, column=col).font = uniform_font
                    ws_far.cell(row=write_row+1, column=col).border = Border()
                mgmt_acct_end_str = ''
                if mgmt_acct_month is not None:
                    mgmt_acct_end_str = mgmt_acct_month.strftime('%b-%Y')
                col_headers = list(combined_df.columns)
                if len(col_headers) >= 2:
                    col_headers[-2] = f"{col_headers[-2]} (as at {mgmt_acct_end_str})"
                    col_headers[-1] = f"{col_headers[-1]} (as at {mgmt_acct_end_str})"
                for col_idx, col_name in enumerate(col_headers, start=1):
                    cell = ws_far.cell(row=write_row+2, column=col_idx, value=col_name)
                    cell.border = header_border
                    cell.font = bold_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                units_row_idx = write_row+3
                units_map = {}
                for col in combined_df.columns:
                    col_lc = str(col).strip().lower()
                    if 'date' in col_lc:
                        units_map[col] = ''
                    elif 'rate' in col_lc:
                        units_map[col] = '%'
                    elif 'depreciation' in col_lc or 'cost' in col_lc or 'addition' in col_lc or 'wdv' in col_lc:
                        units_map[col] = '£'
                    elif 'dep' in col_lc and any(mon in col_lc for mon in ['jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec']):
                        units_map[col] = '£'
                    else:
                        units_map[col] = ''
                for col_idx, col_name in enumerate(combined_df.columns, start=1):
                    cell = ws_far.cell(row=units_row_idx, column=col_idx, value=units_map.get(col_name, ''))
                    cell.font = uniform_font
                    cell.alignment = center_align
                    cell.border = header_border
                # Write data rows, and apply row-wise SUM formulas for Total Depreciation and WDV
                dep_col_idx = None
                wdv_col_idx = None
                acc_dep_col_idx = None
                total_cost_col_idx = None
                # Find column indices for relevant columns
                for idx, col_name in enumerate(combined_df.columns):
                    col_clean = str(col_name).strip().lower().replace(' ', '')
                    if col_clean == 'totaldepreciation':
                        dep_col_idx = idx
                    if col_clean == 'wdv':
                        wdv_col_idx = idx
                    if col_clean == 'accumulateddepreciation':
                        acc_dep_col_idx = idx
                    if col_clean == 'totalcost':
                        total_cost_col_idx = idx
                month_col_indices = [i for i, col in enumerate(combined_df.columns) if any(mon in str(col).lower() for mon in ['jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec'])]
                for r in range(len(combined_df)):
                    for c in range(len(combined_df.columns)):
                        val = combined_df.iloc[r, c]
                        if hasattr(val, 'item'):
                            val = val.item()
                        cell = ws_far.cell(row=write_row+4+r, column=c+1)
                        col_name = combined_df.columns[c]
                        col_lc = str(col_name).strip().lower()
                        # Row-wise SUM for Total Depreciation
                        if dep_col_idx is not None and c == dep_col_idx and acc_dep_col_idx is not None and month_col_indices:
                            acc_dep_cell = ws_far.cell(row=write_row+4+r, column=acc_dep_col_idx+1).coordinate
                            month_cells = [ws_far.cell(row=write_row+4+r, column=idx+1).coordinate for idx in month_col_indices]
                            cell.value = f"=SUM({acc_dep_cell},{','.join(month_cells)})"
                        # Row-wise SUM for WDV
                        elif wdv_col_idx is not None and c == wdv_col_idx and total_cost_col_idx is not None and dep_col_idx is not None:
                            total_cost_cell = ws_far.cell(row=write_row+4+r, column=total_cost_col_idx+1).coordinate
                            dep_cell = ws_far.cell(row=write_row+4+r, column=dep_col_idx+1).coordinate
                            cell.value = f"={total_cost_cell}-{dep_cell}"
                        else:
                            cell.value = val
                        # Add vertical borders to data cells
                        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'))
                        cell.font = uniform_font
                        if c == 0:
                            cell.alignment = date_align
                            if pd.notnull(val):
                                try:
                                    cell.number_format = 'dd-mm-yyyy'
                                except Exception:
                                    pass
                        elif 'rate' in col_lc:
                            cell.alignment = right_align
                            try:
                                if pd.notnull(val):
                                    cell.value = f"{float(val):.0f}%"
                                    cell.number_format = '0%'
                            except Exception:
                                pass
                        elif isinstance(val, (int, float)):
                            cell.alignment = right_align
                            try:
                                if pd.notnull(val):
                                    cell.number_format = '#,##0.00_);(#,##0.00);"-"??_);_(@_)'
                            except Exception:
                                pass
                        elif 'details' in col_lc:
                            cell.alignment = left_align
                        else:
                            cell.alignment = left_align
                total_row = [''] * len(combined_df.columns)
                total_row[1] = 'Total'
                data_for_total = combined_df.copy()
                data_for_total = data_for_total[data_for_total.iloc[:,1].astype(str).str.lower() != 'total']
                data_for_total = data_for_total.dropna(how='all')
                for idx, col in enumerate(combined_df.columns):
                    col_lc = str(col).strip().lower()
                    if 'rate' in col_lc:
                        total_row[idx] = ''
                        continue
                    try:
                        col_data = data_for_total[col]
                        if isinstance(col_data, pd.DataFrame):
                            col_data = col_data.iloc[:,0]
                        if isinstance(col_data, (pd.Series, list, tuple)):
                            col_numeric = pd.to_numeric(col_data, errors='coerce')
                            if pd.api.types.is_numeric_dtype(col_numeric):
                                total_row[idx] = col_numeric.sum(skipna=True)
                    except Exception as e:
                        pass
                total_row_idx = write_row + 4 + len(combined_df) - 1 + 1
                for c in range(len(total_row)):
                    cell = ws_far.cell(row=total_row_idx, column=c+1)
                    col_name = str(combined_df.columns[c]) if c < len(combined_df.columns) else ''
                    col_name_clean = col_name.strip().lower().replace(' ', '')
                    if c == 0:
                        cell.value = None
                        cell.alignment = date_align
                    elif c == 1:
                        cell.value = 'Total'
                        cell.alignment = left_align
                    elif 'rate' in col_name_clean:
                        cell.value = ''
                        cell.alignment = right_align
                    elif col_name_clean in ['totaldepreciation', 'wdv']:
                        col_letter = openpyxl.utils.get_column_letter(c+1)
                        data_start_row = write_row + 4
                        data_end_row = data_start_row + len(combined_df) - 1
                        cell.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
                        cell.alignment = right_align
                    else:
                        col_letter = openpyxl.utils.get_column_letter(c+1)
                        data_start_row = write_row + 4
                        data_end_row = data_start_row + len(combined_df) - 1
                        cell.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
                        cell.alignment = right_align
                    cell.border = header_border
                    cell.font = bold_font
                    if c == 0:
                        try:
                            cell.number_format = 'dd-mm-yyyy'
                        except Exception:
                            pass
                    elif c > 0:
                        try:
                            if not ('rate' in col_name_clean or c == 1):
                                cell.number_format = '#,##0.00_);(#,##0.00);"-"??_);_(@_)'
                        except Exception:
                            pass
                # Auto-adjust column widths for this table in FAR
                for col_idx, col_name in enumerate(combined_df.columns, start=1):
                    max_length = len(str(col_name))
                    for row_offset in range(0, len(combined_df) + 4):
                        cell = ws_far.cell(row=write_row + 2 + row_offset, column=col_idx)
                        val = cell.value
                        if val is not None:
                            val_str = str(val)
                            if len(val_str) > max_length:
                                max_length = len(val_str)
                    ws_far.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_length + 2
                write_row = total_row_idx+3

        # --- Auto-adjust column widths for all tables in all sheets (except excluded) ---
        excluded_sheets = ['FAR', 'Mappings', 'Account Transactions', 'Sheet1', 'Sheet2', 'Sheet3']
        for ws in wb.worksheets:
            if ws.title in excluded_sheets:
                continue
            # Find the header row (assume always at row 15 as per your logic)
            header_row = 15
            # Find the number of columns by checking for non-empty cells in header
            max_col = 0
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=header_row, column=col).value:
                    max_col = col
            if max_col == 0:
                continue
            # For each column, find the max length among header, units, and all data rows
            for col_idx in range(1, max_col + 1):
                max_length = 0
                # Header
                val = ws.cell(row=header_row, column=col_idx).value
                if val is not None:
                    max_length = len(str(val))
                # Units row
                val = ws.cell(row=header_row + 1, column=col_idx).value
                if val is not None and len(str(val)) > max_length:
                    max_length = len(str(val))
                # Data rows (assume data starts at header_row+2 and goes until first empty in col 2)
                data_row = header_row + 2
                while True:
                    val = ws.cell(row=data_row, column=2).value
                    if val is None or str(val).strip() == '':
                        break
                    val = ws.cell(row=data_row, column=col_idx).value
                    if val is not None and len(str(val)) > max_length:
                        max_length = len(str(val))
                    data_row += 1
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_length + 2





    # Apply formatting and adjust column widths after each FormatN routine
    from openpyxl.styles import PatternFill, Font
    excludeSheets = ["Account Transactions", "P&L", "Corporation Tax", "Mappings", "FAR"]
    for ws in wb.worksheets:
        if ws.title not in excludeSheets:
            # Find the last row in the summary table (starting from A15)
            max_row = ws.max_row
            lastRow = None
            for r in range(max_row, 14, -1):
                val = ws.cell(row=r, column=1).value
                if val is not None and str(val).strip() != "":
                    lastRow = r
                    break
            if lastRow is not None and lastRow >= 15:
                # Find the last used column in lastRow
                max_col = ws.max_column
                lastCol = None
                for c in range(max_col, 0, -1):
                    val = ws.cell(row=lastRow, column=c).value
                    if val is not None and str(val).strip() != "":
                        lastCol = c
                        break
                summaryValue = None
                rng = None
                # Loop through columns in lastRow to find the last numeric value
                for i in range(1, (lastCol or 0) + 1):
                    cell = ws.cell(row=lastRow, column=i)
                    val = cell.value
                    if val is not None and isinstance(val, (int, float)) and str(val).strip() != "":
                        summaryValue = val
                        rng = cell
                # Check if C8 is empty before updating it
                c8 = ws.cell(row=8, column=3)
                if (c8.value is None or str(c8.value).strip() == ""):
                    if summaryValue is not None:
                        # Highlight the summary value cell
                        if rng is not None:
                            rng.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        # Copy the summary value to C8 and format conditionally
                        c8.value = summaryValue
                        if summaryValue < 0:
                            c8.number_format = "#,##0.00;(#,##0.00)"
                        elif summaryValue == 0:
                            c8.number_format = "#,##0.00;-#,##0.00;-"
                        else:
                            c8.number_format = "#,##0.00;(#,##0.00)"
                    else:
                        c8.value = "No numeric summary value"
                # --- Adjust column widths for summary table ---
                for col_idx in range(1, (lastCol or 0) + 1):
                    max_length = 0
                    # Check header (row 15)
                    val = ws.cell(row=15, column=col_idx).value
                    if val is not None:
                        max_length = len(str(val))
                    # Check units row (row 16)
                    val = ws.cell(row=16, column=col_idx).value
                    if val is not None and len(str(val)) > max_length:
                        max_length = len(str(val))
                    # Check all data rows
                    for data_row in range(17, lastRow + 1):
                        val = ws.cell(row=data_row, column=col_idx).value
                        if val is not None and len(str(val)) > max_length:
                            max_length = len(str(val))
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_length + 2
                
                # --- Make the last row bold ---
                if lastRow is not None and lastCol is not None:
                    bold_font = Font(name="Gill Sans MT", size=12, bold=True)
                    for col_idx in range(1, lastCol + 1):
                        cell = ws.cell(row=lastRow, column=col_idx)
                        cell.font = bold_font
            else:
                ws.cell(row=8, column=3).value = "No data found"

    # --- MODULE 5: DeleteSheetsWithNoData ---
    ws_to_delete = []
    for ws in wb.worksheets:
            c8 = ws.cell(row=8, column=3).value
            if c8 == "No data found":
                ws_to_delete.append(ws.title)
    for sheet_name in ws_to_delete:
            del wb[sheet_name]
    if ws_to_delete:
            print(f"Deleted sheets with 'No data found' in C8: {', '.join(ws_to_delete)}")

        # 4. Save the modified workbook to output folder and return
    safe_name = re.sub(r'[^A-Za-z0-9._-]', '_', str(output_base_name))
    far_output_name = f'{safe_name}.xlsx'
    output_path = os.path.join(output_folder, far_output_name)
    
    # Save to file path first
    wb.save(output_path)
    print(f"Done! Output saved to: {output_path}")

    # Create a fresh output buffer for return
    output = BytesIO()
    try:
        # Create a new workbook instance from the saved file to avoid stream issues
        with open(output_path, 'rb') as f:
            output.write(f.read())
        output.seek(0)
        return output
    except Exception as e:
        print(f"⚠️ Warning: Error creating output buffer: {e}")
        # Fallback: try to save the workbook directly to buffer
        output.seek(0)
        wb.save(output)
        output.seek(0)
        return output

import streamlit as st

# --- Custom CSS for full-width layout and bigger logo ---
st.markdown("""
    <style>
    .main, .stApp {
        background-color: #f7f9fa;
        width: 100%;
        min-height: 100vh;
        margin: 0;
        padding: 0;
    }
    .header-bar {
    background: linear-gradient(90deg, #003366 0%, #005fa3 100%);
    color: white;
    padding: 1.2rem 0 0.7rem 0;
    margin-bottom: 1.5rem;
    border-radius: 0 0 18px 18px;
    text-align: center;
    box-shadow: 0 2px 8px rgba(0,0,0,0.07);
    width: 100%;         /* Full viewport width */
    margin-left: 0;
    margin-right: 0;
    box-sizing: border-box;
}
    .card {
        background: white;
        border-radius: 16px;
        box-shadow: 0 2px 12px rgba(0,0,0,0.08);
        padding: 2rem 2.5rem 2rem 2.5rem;
        margin: 2rem auto;
        max-width: 1200px; /* increase width */
        width: 95vw; /* fill viewport */
    }
    .centered {
        display: flex;
        flex-direction: column;
        align-items: center;
        justify-content: center;
    }
    .top-left-logo {
        position: fixed;
        top: 18px;
        left: 18px;
        z-index: 1000;
        background: transparent;
    }
    </style>
""", unsafe_allow_html=True)


# --- Header Bar with Logo and Title ---
st.markdown('<div class="top-left-logo">', unsafe_allow_html=True)
st.image("Corient_Logo-01.jpg", width=160)  # Bigger logo
st.markdown('</div>', unsafe_allow_html=True)

st.markdown("""
    <div class="header-bar">
        <h1 style="margin-bottom:0.2rem; font-family:Gill Sans MT,Arial,sans-serif; font-size:2.5rem;">INNControl Processor</h1>
        
    </div>
""", unsafe_allow_html=True)


# --- Card-style Upload Section ---
with st.container():
    upload_card = """
    <style>
    .card.centered {background: white; border-radius: 16px; box-shadow: 0 2px 12px rgba(0,0,0,0.08); padding: 2rem 2.5rem 2rem 2.5rem; margin: 0 auto 2rem auto; max-width: 500px; display: flex; flex-direction: column; align-items: center;}
    </style>
    <div class="card centered">
    """
    st.markdown(upload_card, unsafe_allow_html=True)


from io import BytesIO
from openpyxl import load_workbook
import pandas as pd
import streamlit as st

uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])

if uploaded_file is not None:
    # ✅ Read file once
    file_content = uploaded_file.read()

    # ✅ Reuse the content for both libraries
    excel_for_openpyxl = BytesIO(file_content)
    excel_for_pandas = BytesIO(file_content)

    # ✅ Load workbook and data with data_only=True to avoid image issues
    try:
        wb = load_workbook(excel_for_openpyxl, data_only=True)
        xls = pd.ExcelFile(excel_for_pandas, engine="openpyxl")

       # for sheet in xls.sheet_names:
        #    df = pd.read_excel(xls, sheet_name=sheet, engine="openpyxl")
         #   st.write(f"Sheet: {sheet}")
         #   st.dataframe(df)

    except Exception as e:
        st.error(f"❌ Processing failed: {str(e)}")
else:
    st.warning("Please upload an excel file to continue.")



    st.markdown('</div>', unsafe_allow_html=True)

if uploaded_file is not None:
    with st.container():
        process_card = """
        <style>
        .card.centered {background: white; border-radius: 16px; box-shadow: 0 2px 12px rgba(0,0,0,0.08); padding: 2rem 2.5rem 2rem 2.5rem; margin: 0 auto 2rem auto; max-width: 500px; display: flex; flex-direction: column; align-items: center;}
        </style>
        <div class="card centered">
        """
        st.markdown(process_card, unsafe_allow_html=True)
        if 'processing' not in st.session_state:
            st.session_state['processing'] = False
        if not st.session_state['processing']:
            if st.button('Start Processing', key='start_processing'):
                st.session_state['processing'] = True
                st.rerun()
        else:
            st.markdown(f"<div style='font-size:1.1rem;'><span style='font-size:1.5rem;'>📄</span> <b>Processing:</b> {uploaded_file.name}</div>", unsafe_allow_html=True)
            with st.spinner("Processing... Please wait."):
                st.image("https://media0.giphy.com/media/v1.Y2lkPTc5MGI3NjExY2Uwbm5xeTQwd3A4eWU2bGU4Ym9pdjF0YWRkZzltcGJyOTE5czJ0ZCZlcD12MV9pbnRlcm5hbF9naWZfYnlfaWQmY3Q9Zw/2bYewTk7K2No1NvcuK/giphy.gif", width=200)
                try:
                   
                    output_bytes = process_far_file(file_content)
                    output_data = output_bytes.getvalue()
                    st.session_state['processing'] = False
                    st.markdown("<div style='font-size:1.1rem; color: #155724;'><span style='font-size:1.5rem;'>✅</span> <b>File processed successfully!</b> Download is ready.</div>", unsafe_allow_html=True)
                    st.download_button(
                        label="⬇️ Download Processed Excel",
                        data=output_data,
                        file_name="MA_Processed.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.session_state['processing'] = False
                    st.markdown(f"<div style='font-size:1.1rem; color: #721c24;'><span style='font-size:1.5rem;'>❌</span> <b>Processing failed:</b> {e}</div>", unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
